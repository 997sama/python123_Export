#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @PyCharm   :2023.3.4.PY-233.14475.56
# @Python    :3.12
# @FileName  :test.py
# @Time      :2024/6/8 下午11:17
# @Author    :997
# @E-mail    :A997sama@outlook.com
# --------------------------------
import re
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import Workbook

options = Options()
# options.add_argument("--headless")
options.add_argument("--disable-gpu")
news = {}
user_key = {'username': "xxxxxxxxx", 'password': "xxxxxxxxx"}


service = Service()  # 确保传入正确的msedgedriver路径
driver = webdriver.Edge(service=service, options=options)
driver.maximize_window()
driver.get("https://python123.io/")
wait = WebDriverWait(driver, 5)  # 显式等待时间增加到20秒

iii = 0


# 列出需要剔除的特殊字符的Unicode编码


def main():
    # 定位并点击登录链接
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='links']/div[1]/div[1]/nav/div/div[2]/div[2]/div[2]/div"))).click()
    # 输入账户和密码并点击登录
    wait.until(EC.presence_of_element_located((By.XPATH,
                                               "//*[@id='links']/div[1]/div[2]/div/div[1]/div[1]/div/div[1]/div/div["
                                               "1]/div[2]/form/div[1]/div[2]/input"))).send_keys(
        user_key['username'])
    wait.until(EC.presence_of_element_located((By.XPATH,
                                               "//*[@id='links']/div[1]/div[2]/div/div[1]/div[1]/div/div[1]/div/div["
                                               "1]/div[2]/form/div[2]/div[2]/input"))).send_keys(
        user_key['password'])
    login_button = wait.until(EC.element_to_be_clickable((By.XPATH,
                                                          "//*[@id='links']/div[1]/div[2]/div/div[1]/div[1]/div/div["
                                                          "1]/div/div[1]/div[2]/div[3]/div/div[1]/div/span")))
    webdriver.ActionChains(driver).move_to_element(login_button).click().perform()

    # 点击课程
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div[1]/div[2]/div[2]/div/div"))).click()

    # 点击任务
    wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[1]/div/ul[2]/li[2]/div"))).click()

    # 按照第1章至第8章的顺序依次保存题目和答案
    xpath_list = [
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[7]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[1]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[2]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[3]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[4]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[5]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[6]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[1]/div[2]/div/div[8]",
        "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[2]/div/div/div/div/div[2]/div[2]/div/div/div"
    ]
    questions = {}
    # questions中存有题型、编号、题目、答案
    for xpath in xpath_list:
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        time.sleep(1)
        # print(driver.find_element(By.CLASS_NAME, 'v-title').text)
        topic = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='group-wrapper']/div/div[3]/div/div[1]/div/div[2]/div[1]/div[2]/div/div[2]/div[1]/b")))
        if topic.text == "单项选择题":
            wait.until(EC.element_to_be_clickable((By.XPATH,
                                                   "//*[@id='group-wrapper']/div/div[3]/div/div[1]/div/div[2]/div["
                                                   "1]/div[2]/div/div[2]/div[2]/div[1]"))).click()
            multiple_choice_questions(questions)
            wait.until(EC.element_to_be_clickable((By.XPATH,
                                                   "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div["
                                                   "1]/div/ul[2]/li[2]/div"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            second_topic = wait.until(EC.presence_of_element_located((By.XPATH,
                                                                      "//*[@id='group-wrapper']/div/div[3]/div/div["
                                                                      "1]/div/div[2]/div[1]/div[2]/div/div[3]/div["
                                                                      "1]/b")))
            if second_topic and second_topic.text == "程序设计题":
                wait.until(EC.element_to_be_clickable((By.XPATH,
                                                       "//*[@id='group-wrapper']/div/div[3]/div/div[1]/div/div["
                                                       "2]/div[1]/div[2]/div/div[3]/div[2]/div[1]"))).click()
                programming_questions(questions)
                wait.until(EC.element_to_be_clickable((By.XPATH,
                                                       "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div["
                                                       "1]/div/ul[2]/li[2]/div"))).click()
                wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        elif topic.text == "程序设计题":
            wait.until(EC.element_to_be_clickable((By.XPATH,
                                                   '//*[@id="group-wrapper"]/div/div[3]/div/div[1]/div/div[2]/div['
                                                   '1]/div[2]/div/div[2]/div[2]/div[1]'))).click()
            programming_questions(questions)
            wait.until(EC.element_to_be_clickable((By.XPATH,
                                                   "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div["
                                                   "1]/div/ul[2]/li[2]/div"))).click()
            wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='app']/div/div[1]/div[1]/div[2]/section/div/div/div[1]/div/ul[2]/li[2]/div"))).click()

    save_excel(questions)
    # 关闭浏览器
    driver.quit()

    return 0


def multiple_choice_questions(questions):
    global iii
    # print("选择题")
    time.sleep(2)

    problem_cards = driver.find_elements(By.CLASS_NAME, 'problem-card')

    for card in problem_cards:
        problem_id = card.find_element(By.CLASS_NAME, 'problem-id').text
        # print(f"编号：{problem_id}")
        # 提取题目内容
        question = card.find_element(By.CLASS_NAME, 'mce-content-body').text
        # 提取选项
        options = card.find_elements(By.CLASS_NAME, 'options')
        options_text = []
        text = ""
        for option in options:
            option_text = option.find_element(By.CLASS_NAME, 'content').text
            options_text.append(option_text)
        # print(f"题目: {remove_invisible_chars(question)}")
        for i, option_text in enumerate(options_text):
            text += f"选项 {chr(65 + i)}: {remove_invisible_chars(option_text)} "
        # 打印答案（标记为 'is-primary' 的选项）
        try:
            answer = card.find_element(By.CLASS_NAME, 'is-primary').text
        except:
            answer = ""
        # print(f"答案: {answer}\n")

        questions[problem_id + " " + str(iii)] = {
            "题型": "选择题",
            "题目": remove_invisible_chars(question) + " " + text,
            "答案": answer,
        }
        iii += 1
    return questions


def programming_questions(questions):
    global iii
    # print("编程题")
    time.sleep(3)

    problem_id = driver.find_element(By.XPATH,
                                     '//*[@id="group-wrapper"]/div/div[3]/div/div[1]/div/div[1]/div[1]/div/div['
                                     '1]/div/div[1]').text
    # print(f"编号：{problem_id}")
    question = driver.find_element(By.CLASS_NAME, 'mce-content-body').text
    # print(f"题目：{remove_invisible_chars(question)}")

    try:
        answer = driver.find_element(By.CSS_SELECTOR, '.ace_content .ace_text-layer').text
    except:
        answer = ""
    # print(f"答案：{remove_invisible_chars(answer)}")
    questions[problem_id + " " + str(iii)] = {
        "题型": "编程题",
        "题目": remove_invisible_chars(question),
        "答案": remove_invisible_chars(answer)
    }
    iii += 1

    # 找到按钮元素
    button = driver.find_element(By.XPATH,
                                 '//*[@id="group-wrapper"]/div/div[3]/div/div[1]/div/div[1]/div[1]/div/div['
                                 '1]/div/div[2]/div[4]')
    # 检查按钮是否包含disabled属性和代表不可激活状态的class
    if button.get_attribute("disabled") == "disabled" or "cursor-not-allowed" in button.get_attribute("class"):
        pass
        # print("按钮处于不可激活状态")
    else:
        button.click()
        programming_questions(questions)
    return questions


def remove_invisible_chars(text):
    # 使用正则表达式匹配所有的不可见字符并移除，还有换行符
    cleaned_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F\u202A-\u202E\n\xb6\xac]', '', text)
    # 剔除掉空格，目前没有更好的想法，剔除掉空格方便搜索
    cleaned_text = re.sub(r' ', '', cleaned_text)
    # cleaned_text = re.sub(r'描述', ' 描述：', cleaned_text)
    # cleaned_text = re.sub(r'输入', '输入：', cleaned_text)
    # cleaned_text = re.sub(r'输出', ' 输出：', cleaned_text)
    cleaned_text = re.sub(r'·', ' ', cleaned_text)
    cleaned_text = re.sub(r'¤', '\n', cleaned_text)
    return cleaned_text


def save_excel(questions):
    # 把questions保存进excel表格
    wb = Workbook()

    # 激活当前工作表
    ws = wb.active

    # 添加字典数据到工作表中
    # 将问题编号和问题内容添加到第一行
    ws['A1'] = '问题编号'
    ws['B1'] = '题型'
    ws['C1'] = '题目'
    ws['D1'] = '答案'
    for question in questions.items():
        # 将问题数据添加到下一行
        problem_id, problem_info = question
        ws.append([problem_id, problem_info['题型'], problem_info['题目'], problem_info['答案']])

    # 保存Excel文件，文件名为用户名+".xlsx"
    file_name = f"{user_key['username']}.xlsx"
    wb.save(file_name)
    print("successful!")
    return 0


if __name__ == "__main__":
    main()
else:
    print(f"Import {__name__} Successful")
